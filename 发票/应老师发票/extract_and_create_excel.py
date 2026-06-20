import fitz
import os
import re
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

folder = r"C:\Users\P1368\Desktop\应发票"
output_file = r"C:\Users\P1368\Desktop\应发票\发票清单.xlsx"

# 分类关键词
type_keywords = {
    '火车票': ['铁路', 'G6', 'G1', 'G8', 'G7', 'D2', 'G7795', 'G885', 'G1057', 'G6953', '退票', '二等座', '曲阜东', '蒙山', '临沂北', '宁海', '西昌西', '甘洛南', '宁波'],
    '保险': ['保险', '中华联合', '车船税'],
    '住宿': ['酒店', '维也纳', '住宿', '房费', '代订房费'],
    '餐饮': ['餐饮', '饭店', '餐费', '本味菜馆'],
    '办公用品': ['打印', '复印', '办公'],
    '交通运输': ['加油', '过路费', '通行费', '停车'],
    '茶叶': ['茶叶', '茶', '贡府'],
    '其他': []
}

def classify_invoice(text):
    for inv_type, keywords in type_keywords.items():
        for kw in keywords:
            if kw.lower() in text.lower():
                return inv_type
    return '其他'

def extract_invoice_info(file_path, file_name):
    try:
        if file_name.lower().endswith('.pdf'):
            doc = fitz.open(file_path)
            text = ""
            for page in doc:
                text += page.get_text()
            doc.close()
        elif file_name.lower().endswith('.ofd'):
            # OFD暂时跳过
            return None
        else:
            return None
        
        # 提取发票号码
        invoice_no_match = re.search(r'发票号码[：:]\s*(\d+)', text)
        invoice_no = invoice_no_match.group(1) if invoice_no_match else ''
        
        # 提取开票日期
        date_match = re.search(r'(\d{4})[年/](\d{2})[月/](\d{2})[日]?', text)
        if date_match:
            invoice_date = f"{date_match.group(1)}-{date_match.group(2)}-{date_match.group(3)}"
        else:
            # 尝试其他格式
            date_match2 = re.search(r'(\d{4})-(\d{2})-(\d{2})', text)
            invoice_date = f"{date_match2.group(1)}-{date_match2.group(2)}-{date_match2.group(3)}" if date_match2 else ''
        
        # 提取购买方名称
        buyer_match = re.search(r'购买方名称[：:]\s*([^\n]+)', text)
        if not buyer_match:
            buyer_match = re.search(r'名\s*称[：:]\s*([^\n]+)', text)
        buyer = buyer_match.group(1).strip() if buyer_match else ''
        
        # 提取销售方
        seller_match = re.search(r'销售方[^\n]*名称[：:]\s*([^\n]+)', text)
        if not seller_match:
            seller_match = re.search(r'销方[^\n]*名\s*称[：:]\s*([^\n]+)', text)
        if not seller_match:
            # 尝试从文件名提取
            if '凉山州五源兴' in file_name:
                seller = '凉山州五源兴农业科技有限公司'
            elif '应红林' in file_name:
                seller = '应红林（个人）'
            elif '飞猪' in file_name or '代订房费' in file_name:
                seller = '飞猪/在线旅游平台'
            else:
                seller = ''
        else:
            seller = seller_match.group(1).strip()
        
        # 提取金额
        amount_match = re.search(r'[金票]额[：:]?\s*[¥￥]?\s*([\d,]+\.?\d*)', text)
        if not amount_match:
            amount_match = re.search(r'¥\s*([\d,]+\.?\d*)', text)
        amount = amount_match.group(1) if amount_match else ''
        
        # 提取行程信息（火车票）
        route_info = ''
        if '铁路' in text or 'G6' in text or 'G1' in text or 'G8' in text or 'G7' in text or 'D2' in text:
            # 提取出发地和目的地
            stations = re.findall(r'([\u4e00-\u9fa5]+站)', text)
            if stations:
                route_info = '→'.join(stations[:2]) if len(stations) >= 2 else stations[0] if stations else ''
            
            # 提取车次
            train_match = re.search(r'([GKD]\d+)[^\d]', text)
            if train_match:
                route_info = f"{train_match.group(1)} {route_info}"
            
            # 提取座位
            seat_match = re.search(r'([一二三]等座|[商务座])', text)
            seat = seat_match.group(1) if seat_match else ''
            
            # 提取票价
            price_match = re.search(r'票\s*价[：:]?\s*[¥￥]?\s*([\d,]+\.?\d*)', text)
            if price_match:
                amount = price_match.group(1)
            
            if '退票' in text:
                route_info += ' [退票]'
        
        # 判断是否退票
        is_refund = '退票' in text
        
        # 分类
        inv_type = classify_invoice(text)
        
        # 备注
        remark = ''
        if is_refund:
            remark = '退票'
        if '免税' in text:
            remark += ' | 免税商品'
        
        return {
            'file_name': file_name,
            'invoice_no': invoice_no,
            'invoice_date': invoice_date,
            'buyer': buyer,
            'seller': seller,
            'amount': amount,
            'type': inv_type,
            'route_info': route_info,
            'remark': remark,
            'is_refund': is_refund
        }
    except Exception as e:
        return {
            'file_name': file_name,
            'invoice_no': '',
            'invoice_date': '',
            'buyer': '',
            'seller': '',
            'amount': '',
            'type': '读取失败',
            'route_info': '',
            'remark': f'错误: {str(e)}',
            'is_refund': False
        }

# 收集所有文件
results = []
for f in sorted(os.listdir(folder)):
    if f.lower().endswith(('.pdf', '.ofd')):
        path = os.path.join(folder, f)
        info = extract_invoice_info(path, f)
        if info:
            results.append(info)

# 创建Excel
wb = Workbook()
ws = wb.active
ws.title = "发票清单"

# 样式
header_fill = PatternFill(start_color="1e40af", end_color="1e40af", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
cell_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# 表头
headers = ['序号', '文件名', '发票号码', '开票日期', '发票类型', '购买方', '销售方', '金额(元)', '行程信息', '是否退票', '备注']
for col, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = cell_alignment
    cell.border = thin_border

# 数据
for row, r in enumerate(results, 2):
    ws.cell(row=row, column=1, value=row-1).border = thin_border
    ws.cell(row=row, column=2, value=r['file_name']).border = thin_border
    ws.cell(row=row, column=3, value=r['invoice_no']).border = thin_border
    ws.cell(row=row, column=4, value=r['invoice_date']).border = thin_border
    ws.cell(row=row, column=5, value=r['type']).border = thin_border
    ws.cell(row=row, column=6, value=r['buyer']).border = thin_border
    ws.cell(row=row, column=7, value=r['seller']).border = thin_border
    
    amt_cell = ws.cell(row=row, column=8, value=r['amount'])
    amt_cell.border = thin_border
    if r['amount']:
        try:
            amt_cell.number_format = '#,##0.00'
        except:
            pass
    
    ws.cell(row=row, column=9, value=r['route_info']).border = thin_border
    ws.cell(row=row, column=10, value='是' if r['is_refund'] else '否').border = thin_border
    ws.cell(row=row, column=11, value=r['remark']).border = thin_border
    
    for col in range(1, 12):
        ws.cell(row=row, column=col).alignment = Alignment(horizontal="center", vertical="center")

# 统计
ws2 = wb.create_sheet("汇总统计")
stats_data = [
    ['发票类型', '数量', '金额合计(元)'],
]
type_stats = {}
for r in results:
    t = r['type']
    if t not in type_stats:
        type_stats[t] = {'count': 0, 'amount': 0}
    type_stats[t]['count'] += 1
    if r['amount']:
        try:
            type_stats[t]['amount'] += float(r['amount'].replace(',', ''))
        except:
            pass

for t, s in sorted(type_stats.items(), key=lambda x: -x[1]['amount']):
    stats_data.append([t, s['count'], f"{s['amount']:,.2f}"])

# 退票统计
refund_count = sum(1 for r in results if r['is_refund'])
refund_amount = sum(float(r['amount'].replace(',','')) for r in results if r['is_refund'] and r['amount'])
stats_data.append(['', '', ''])
stats_data.append(['退票小计', refund_count, f"{refund_amount:,.2f}"])

# 有效发票统计
valid_count = len(results) - refund_count
valid_amount = sum(float(r['amount'].replace(',','')) for r in results if not r['is_refund'] and r['amount'])
stats_data.append(['', '', ''])
stats_data.append(['有效发票合计', valid_count, f"{valid_amount:,.2f}"])
stats_data.append(['全部发票合计', len(results), f"{sum(float(r['amount'].replace(',','')) for r in results if r['amount'] and not r['is_refund']):,.2f}"])

for row, data in enumerate(stats_data, 1):
    for col, val in enumerate(data, 1):
        cell = ws2.cell(row=row, column=col, value=val)
        cell.border = thin_border
        if row == 1:
            cell.fill = header_fill
            cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

# 列宽
ws.column_dimensions['A'].width = 6
ws.column_dimensions['B'].width = 45
ws.column_dimensions['C'].width = 20
ws.column_dimensions['D'].width = 12
ws.column_dimensions['E'].width = 10
ws.column_dimensions['F'].width = 30
ws.column_dimensions['G'].width = 25
ws.column_dimensions['H'].width = 12
ws.column_dimensions['I'].width = 25
ws.column_dimensions['J'].width = 8
ws.column_dimensions['K'].width = 20

ws2.column_dimensions['A'].width = 15
ws2.column_dimensions['B'].width = 10
ws2.column_dimensions['C'].width = 18

wb.save(output_file)
print(f"完成！共处理 {len(results)} 个文件")
print(f"Excel已保存到: {output_file}")
