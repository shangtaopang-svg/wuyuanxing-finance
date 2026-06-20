#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import json, re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

with open(r"C:\Users\P1368\Desktop\甘洛项目\发票\党参\发票汇总.json", encoding="utf-8") as f:
    data = json.load(f)

# 从 _raw 中精确提取销售方姓名（开票人）
def get_seller_name(raw):
    m = re.search(r'销\s*名称[：:\s]*([^\s\n]{2,6})', raw)
    if m:
        return m.group(1).strip()
    m = re.search(r'开票人[：:\s]*([^\s\n]{2,6})', raw)
    if m:
        return m.group(1).strip()
    return ""

def get_qty_price(raw):
    # 数量 单价
    m = re.search(r'(\d+(?:\.\d+)?)\s+(\d+(?:\.\d+)?)\s+\d+(?:\.\d+)?\s+免税', raw)
    if m:
        return m.group(1), m.group(2)
    return "", ""

rows = []
for item in data:
    raw = item.get("_raw", "")
    seller = get_seller_name(raw)
    qty, price = get_qty_price(raw)
    amount_str = item["金额"].replace("¥", "").replace(",", "").strip()
    try:
        amount = float(amount_str)
    except:
        amount = 0.0
    rows.append({
        "序号": item["序号"],
        "姓名": seller,
        "发票号码": item["发票号码"],
        "开票日期": item["开票日期"],
        "货物名称": "党参苗",
        "数量(kg)": qty,
        "单价(元/kg)": price,
        "金额(元)": amount,
        "文件名": item["文件名"]
    })

wb = Workbook()
ws = wb.active
ws.title = "党参发票汇总"

# 标题行
title_font = Font(name="微软雅黑", bold=True, size=11, color="FFFFFF")
header_fill = PatternFill("solid", start_color="1F4E79")
data_fill1  = PatternFill("solid", start_color="FFFFFF")
data_fill2  = PatternFill("solid", start_color="EBF3FB")
money_fill  = PatternFill("solid", start_color="FFF2CC")
total_fill  = PatternFill("solid", start_color="D6E4F0")
border_side = Side(style="thin", color="BFBFBF")
thin_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
center = Alignment(horizontal="center", vertical="center")
left   = Alignment(horizontal="left",   vertical="center")
right  = Alignment(horizontal="right",  vertical="center")

# 合并标题
ws.merge_cells("A1:I1")
ws["A1"] = "甘洛项目·党参苗采购发票汇总表"
ws["A1"].font = Font(name="微软雅黑", bold=True, size=14, color="1F4E79")
ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws["A1"].fill = PatternFill("solid", start_color="D6E4F0")
ws.row_dimensions[1].height = 32

ws.merge_cells("A2:I2")
ws["A2"] = "购买方：凉山州五源兴农业科技有限公司    开票日期：2026年03月23日    共39份发票"
ws["A2"].font = Font(name="微软雅黑", size=10, color="595959")
ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
ws["A2"].fill = PatternFill("solid", start_color="EBF3FB")
ws.row_dimensions[2].height = 20

headers = ["序号", "销售方姓名", "发票号码", "开票日期", "货物名称", "数量(kg)", "单价(元/kg)", "金额(元)", "来源文件"]
col_widths = [6, 12, 24, 16, 10, 10, 12, 14, 28]

for col_i, (h, w) in enumerate(zip(headers, col_widths), 1):
    cell = ws.cell(row=3, column=col_i, value=h)
    cell.font = title_font
    cell.fill = header_fill
    cell.alignment = center
    cell.border = thin_border
    ws.column_dimensions[get_column_letter(col_i)].width = w
ws.row_dimensions[3].height = 22

# 数据行
for i, row in enumerate(rows, 4):
    fill = data_fill2 if (i % 2 == 0) else data_fill1
    vals = [
        row["序号"], row["姓名"], row["发票号码"], row["开票日期"],
        row["货物名称"], row["数量(kg)"], row["单价(元/kg)"], row["金额(元)"], row["文件名"]
    ]
    aligns = [center, center, left, center, center, center, center, right, left]
    for col_i, (val, aln) in enumerate(zip(vals, aligns), 1):
        cell = ws.cell(row=i, column=col_i, value=val)
        cell.font = Font(name="微软雅黑", size=10)
        cell.fill = money_fill if col_i == 8 else fill
        cell.alignment = aln
        cell.border = thin_border
        if col_i == 8:
            cell.number_format = '#,##0.00'
    ws.row_dimensions[i].height = 18

# 合计行
total_row = 4 + len(rows)
ws.merge_cells(f"A{total_row}:G{total_row}")
ws[f"A{total_row}"] = f"合  计（共 {len(rows)} 份）"
ws[f"A{total_row}"].font = Font(name="微软雅黑", bold=True, size=11, color="1F4E79")
ws[f"A{total_row}"].alignment = center
ws[f"A{total_row}"].fill = total_fill
ws[f"A{total_row}"].border = thin_border

total_cell = ws.cell(row=total_row, column=8, value=f"=SUM(H4:H{total_row-1})")
total_cell.font = Font(name="微软雅黑", bold=True, size=11, color="C00000")
total_cell.fill = total_fill
total_cell.alignment = right
total_cell.border = thin_border
total_cell.number_format = '#,##0.00'

ws.cell(row=total_row, column=9).fill = total_fill
ws.cell(row=total_row, column=9).border = thin_border
ws.row_dimensions[total_row].height = 24

# 冻结首行(含标题)
ws.freeze_panes = "A4"

# 统计分页 Sheet2：按姓名汇总
ws2 = wb.create_sheet("按姓名汇总")
ws2.merge_cells("A1:E1")
ws2["A1"] = "按销售方姓名汇总"
ws2["A1"].font = Font(name="微软雅黑", bold=True, size=13, color="1F4E79")
ws2["A1"].alignment = center
ws2["A1"].fill = PatternFill("solid", start_color="D6E4F0")
ws2.row_dimensions[1].height = 28

h2 = ["姓名", "发票数量", "合计金额(元)", "占比", "备注"]
w2 = [14, 12, 18, 10, 30]
for ci, (h, w) in enumerate(zip(h2, w2), 1):
    cell = ws2.cell(row=2, column=ci, value=h)
    cell.font = title_font
    cell.fill = header_fill
    cell.alignment = center
    cell.border = thin_border
    ws2.column_dimensions[get_column_letter(ci)].width = w
ws2.row_dimensions[2].height = 22

from collections import defaultdict
summary = defaultdict(lambda: {"count": 0, "total": 0.0})
for row in rows:
    summary[row["姓名"]]["count"] += 1
    summary[row["姓名"]]["total"] += row["金额(元)"]

grand_total = sum(v["total"] for v in summary.values())
color_map = {"谢想红": "FCE4D6", "杨德魁": "E2EFDA", "杨小军": "DDEBF7", "王秀英": "FFF2CC", "闫长毛": "EDE7F6"}
sr = 3
for name, v in sorted(summary.items(), key=lambda x: -x[1]["total"]):
    bg = color_map.get(name, "F5F5F5")
    row_fill = PatternFill("solid", start_color=bg)
    data = [name, v["count"], v["total"], v["total"]/grand_total if grand_total else 0, ""]
    afmts = [center, center, right, center, left]
    for ci, (val, af) in enumerate(zip(data, afmts), 1):
        cell = ws2.cell(row=sr, column=ci, value=val)
        cell.font = Font(name="微软雅黑", size=10)
        cell.fill = row_fill
        cell.alignment = af
        cell.border = thin_border
        if ci == 3: cell.number_format = '#,##0.00'
        if ci == 4: cell.number_format = '0.0%'
    ws2.row_dimensions[sr].height = 20
    sr += 1

# 合计
ws2.merge_cells(f"A{sr}:B{sr}")
ws2[f"A{sr}"] = f"合计（{len(summary)}人）"
ws2[f"A{sr}"].font = Font(name="微软雅黑", bold=True, size=11)
ws2[f"A{sr}"].alignment = center
ws2[f"A{sr}"].fill = total_fill
ws2[f"A{sr}"].border = thin_border
tc = ws2.cell(row=sr, column=3, value=f"=SUM(C3:C{sr-1})")
tc.font = Font(name="微软雅黑", bold=True, size=11, color="C00000")
tc.fill = total_fill; tc.alignment = right; tc.border = thin_border; tc.number_format = '#,##0.00'
for ci in [4,5]:
    c = ws2.cell(row=sr, column=ci)
    c.fill = total_fill; c.border = thin_border

out = r"C:\Users\P1368\Desktop\甘洛项目\发票\党参\党参苗发票汇总.xlsx"
wb.save(out)
print(f"Excel已保存：{out}")
print(f"共{len(rows)}条记录，合计金额：¥{grand_total:,.2f}")
